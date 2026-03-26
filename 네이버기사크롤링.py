from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from openpyxl import Workbook

try:
    # Chrome 옵션 설정 (headless 모드 제거)
    chrome_options = Options()
    # chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    # Chrome 드라이버 설정 (자동 다운로드)
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

    # 네이버 검색 결과 URL
    url = "https://search.naver.com/search.naver?ie=UTF-8&query=%EB%B0%98%EB%8F%84%EC%B2%B4&sm=chr_hty"

    # 웹페이지 로드
    driver.get(url)

    # 페이지 로드 대기
    time.sleep(5)

    # 뉴스 제목 추출 (클래스 'news_tit'인 a 태그 찾기)
    news_titles = driver.find_elements(By.CLASS_NAME, "news_tit")

    # 비디오 제목 추출 (비디오 섹션의 제목 span 찾기)
    video_titles = driver.find_elements(By.CSS_SELECTOR, "span.sds-comps-text-ellipsis-1.sds-comps-text-type-headline1")

    # 브라우저 닫기
    driver.quit()

    print(f"뉴스 제목 개수: {len(news_titles)}")
    print(f"비디오 제목 개수: {len(video_titles)}")

    # 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "크롤링 결과"

    # 헤더 추가
    ws['A1'] = "신문기사 제목"
    ws['B1'] = "비디오 제목"

    # 뉴스 제목 저장
    for i, title in enumerate(news_titles, start=2):
        ws[f'A{i}'] = title.text

    # 비디오 제목 저장
    for i, title in enumerate(video_titles, start=2):
        ws[f'B{i}'] = title.text

    # 파일 저장
    wb.save("naver_result.xlsx")

    print("크롤링 결과가 naver_result.xlsx 파일에 저장되었습니다.")

except Exception as e:
    print(f"오류 발생: {e}")
    import traceback
    traceback.print_exc()
